import sys
sys.setrecursionlimit(5000)

import pandas as pd
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def process_food_data():
    # 隱藏 tkinter 主視窗
    root = tk.Tk()
    root.withdraw()
    
    # 1. 選擇檔案與資料夾
    print("請在彈出的視窗中選擇「原始用餐紀錄」檔案...")
    input_file = filedialog.askopenfilename(
        title="請選擇原始用餐紀錄檔案",
        filetypes=[("Excel/CSV 檔案", "*.xlsx *.xls *.csv")]
    )
    if not input_file: return
        
    print("請選擇要儲存「分類後檔案」的資料夾...")
    output_folder = filedialog.askdirectory(title="請選擇儲存資料夾")
    if not output_folder: return

    # 讀取資料
    try:
        if input_file.endswith('.csv'):
            df = pd.read_csv(input_file)
        else:
            df = pd.read_excel(input_file)
    except Exception as e:
        print(f"讀取檔案失敗: {e}")
        return

    # 2. 處理日期與民國年前綴
    try:
        date_series = pd.to_datetime(df['日期']).dropna()
        first_valid_date = date_series.iloc[0]
        roc_year = first_valid_date.year - 1911
        month_str = f"{first_valid_date.month:02d}"
        date_prefix = f"{roc_year}{month_str}"
        df['日期'] = pd.to_datetime(df['日期']).dt.strftime('%Y-%m-%d')
    except Exception as e:
        date_prefix = "00000"
        roc_year, month_str = "未知", "未知"

    # 3. 判斷計費
    df['有效餐數'] = df[['訂購數量', '刷卡數量']].max(axis=1)
    df['計費'] = df['有效餐數'] * 40

    departments = df['單位'].dropna().unique()
    print(f"成功讀取資料，預計產生 {len(departments)} 個單位的檔案。加入精美簽核表格中...")

    # 4. 依單位拆分檔案
    for dept in departments:
        dept_df = df[df['單位'] == dept].copy()
        
        detail_df = dept_df.drop(columns=['有效餐數'])
        
        summary_df = dept_df.groupby(['帳號', '姓名', '單位'], as_index=False).agg(
            刷卡數量=('有效餐數', 'sum'),
            計費=('計費', 'sum')
        )
        total_meals = summary_df['刷卡數量'].sum()
        total_cost = total_meals * 40 
        
        total_row = pd.DataFrame({
            '帳號': [''], '姓名': [''], '單位': [''], '刷卡數量': [total_meals], '計費': ['']
        })
        summary_df = pd.concat([summary_df, total_row], ignore_index=True)
        
        safe_dept_name = str(dept).replace("/", "_").replace("\\", "_")
        filename = f"{date_prefix}供餐系統收費表-{safe_dept_name}.xlsx"
        filepath = os.path.join(output_folder, filename)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            detail_df.to_excel(writer, sheet_name=safe_dept_name, index=False)
            summary_df.to_excel(writer, sheet_name='統計次數', index=False)
            
            header_font = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
            data_font = Font(name="微軟正黑體", size=11)
            center_align = Alignment(horizontal="center", vertical="center")
            thin_side = Side(border_style="thin", color="D9D9D9")
            thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                worksheet.print_title_rows = '1:1'
                worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
                worksheet.sheet_properties.pageSetUpPr.fitToPage = True
                worksheet.page_setup.fitToWidth = 1
                worksheet.page_setup.fitToHeight = False
                worksheet.print_options.horizontalCentered = True
                
                # 填入樣式
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=1), start=1):
                    for cell in row:
                        if cell.value is not None or row_idx <= worksheet.max_row: 
                            cell.border = thin_border
                        if row_idx == 1:
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = center_align
                        else:
                            cell.font = data_font
                            if isinstance(cell.value, (int, float)):
                                cell.alignment = Alignment(horizontal="right", vertical="center")
                            else:
                                cell.alignment = center_align
                                
                # 自動計算欄寬
                for col in worksheet.columns:
                    max_len = 0
                    for cell in col:
                        if cell.value is not None:
                            val_str = str(cell.value)
                            str_len = sum(2 if ord(char) > 127 else 1 for char in val_str)
                            if str_len > max_len:
                                max_len = str_len
                    col_letter = col[0].column_letter
                    worksheet.column_dimensions[col_letter].width = max(max_len + 5, 12)

            # 🔥 【位置修正：將黃底表格改寫入第一頁 (明細頁)】
            detail_ws = writer.sheets[safe_dept_name]
            start_row = detail_ws.max_row + 3 # 空兩行後開始寫
            
            # 定義表格專屬樣式
            sign_font = Font(name="微軟正黑體", size=11, bold=True)
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # 正黃色
            black_border = Border(
                left=Side(style='thin', color='000000'), 
                right=Side(style='thin', color='000000'), 
                top=Side(style='thin', color='000000'), 
                bottom=Side(style='thin', color='000000')
            ) # 實體黑框線
            
            # 因為明細表比較寬，我們將總結表對齊在最右邊的 G欄 (第7欄) 與 H欄 (第8欄)
            
            # 1. 第一列表頭 (黃底，合併 G 與 H)
            detail_ws.merge_cells(start_row=start_row, start_column=7, end_row=start_row, end_column=8)
            detail_ws.cell(row=start_row, column=7, value=f"{roc_year}年{month_str}月 月份總結")
            for col in (7, 8):
                c = detail_ws.cell(row=start_row, column=col)
                c.font = sign_font
                c.fill = yellow_fill
                c.border = black_border
                c.alignment = center_align
                
            # 2. 第二列單位 (白底，合併 G 與 H)
            row_dept = start_row + 1
            detail_ws.merge_cells(start_row=row_dept, start_column=7, end_row=row_dept, end_column=8)
            detail_ws.cell(row=row_dept, column=7, value=f"單位：{dept}")
            for col in (7, 8):
                c = detail_ws.cell(row=row_dept, column=col)
                c.font = sign_font
                c.border = black_border
                c.alignment = center_align

            # 3. 第三列總筆數 (黃底，G欄標題、H欄數值)
            row_count = start_row + 2
            detail_ws.cell(row=row_count, column=7, value="總筆數")
            detail_ws.cell(row=row_count, column=8, value=f"{total_meals} 筆")
            for col in (7, 8):
                c = detail_ws.cell(row=row_count, column=col)
                c.font = sign_font
                c.fill = yellow_fill
                c.border = black_border
                c.alignment = center_align

            # 4. 第四列總金額 (黃底，G欄標題、H欄數值)
            row_cost = start_row + 3
            detail_ws.cell(row=row_cost, column=7, value="總金額")
            detail_ws.cell(row=row_cost, column=8, value=f"{total_cost} 元")
            for col in (7, 8):
                c = detail_ws.cell(row=row_cost, column=col)
                c.font = sign_font
                c.fill = yellow_fill
                c.border = black_border
                c.alignment = center_align

            # 5. 第六列簽名區 (空一行後開始)
            # 🔥 關鍵修正：橫跨 D, E, F, G, H (第4到第8欄)，給予極大的空間確保不裁切！
            row_sign = start_row + 5
            detail_ws.merge_cells(start_row=row_sign, start_column=4, end_row=row_sign, end_column=8)
            c_sign = detail_ws.cell(row=row_sign, column=4, value="伙食管理委員會審核：_________________________")
            c_sign.font = sign_font
            c_sign.alignment = Alignment(horizontal="right", vertical="center")

    print("-" * 30)
    print(f"所有檔案處理完成！總結表格已移至第一頁右下角，並加寬簽名區塊。\n儲存路徑:{output_folder}")

if __name__ == '__main__':
    process_food_data()
